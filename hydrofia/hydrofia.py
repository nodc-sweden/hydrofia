import csv
import datetime
import pathlib
import re
import sys
from typing import Protocol

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from hydrofia import utils

# if typing.TYPE_CHECKING:
#     from hydrofia.calculate import HydrofiaData

if getattr(sys, 'frozen', False):
    DIRECTORY = pathlib.Path(sys.executable).parent
else:
    DIRECTORY = pathlib.Path(__file__).parent


class HydrofiaExportFileDiscrete:

    SAMPLE_NAME_PATTERNS = [
        re.compile('^{}{}-{}{}{}$'.format(
            r'(?P<year>\d{4})',
            r'(?P<ship>\d{4})',
            r'(?P<serno>\d{4})',
            r'(?P<sep>[-_])',
            r'(?P<depth>\d*)'
        )),
        re.compile('^{}{}-{}{}{}$'.format(
            r'(?P<year>\d{2})',
            r'(?P<ship>\d{4})',
            r'(?P<serno>\d{4})',
            r'(?P<sep>[-_])',
            r'(?P<depth>\d*)'
        )),
        re.compile('^{}{}-{}{}{}$'.format(
            r'(?P<year>\d{4})',
            r'(?P<ship>\d{4})',
            r'(?P<serno>\d{4})',
            r'(?P<sep>[-_])',
            r'(?P<depth>[xX])'
        )),
        re.compile('^{}{}-{}{}{}$'.format(
            r'(?P<year>\d{2})',
            r'(?P<ship>\d{4})',
            r'(?P<serno>\d{4})',
            r'(?P<sep>[-_])',
            r'(?P<depth>[xX])'
        ))
    ]

    def __init__(self, path: str | pathlib.Path) -> None:
        self.path = pathlib.Path(path)

        self._header_original: list[str] | None = None
        self._units_original: list[str] | None = None
        self._data: pd.DataFrame | None = None

        self._load_file()
        self._filter_data()
        self._add_columns()
        self._reorder_columns()

    def __str__(self) -> str:
        return '\n'.join([
            f'{self.__class__.__name__}: {self.path}',
            f'  Nr data lines: {len(self.data)}',
            f'  Start time   : {min(self.data["timestamp"])}',
            f'  End time     : {max(self.data["timestamp"])}',
        ])

    def __getitem__(self, item: str) -> str:
        return self.data[item]

    @property
    def data(self) -> pd.DataFrame:
        return self._data

    @property
    def columns(self) -> list[str]:
        return list(self.data.columns)

    def get_data(self) -> pd.DataFrame:
        return self.data

    def filter_data_by_date(self, start_date: datetime.date = None, end_date: datetime.date = None) -> None:
        filter_boolean = np.full(len(self.data), True)
        if start_date:
            filter_boolean = filter_boolean & (self.data['date'] >= start_date)
        if end_date:
            filter_boolean = filter_boolean & (self.data['date'] <= end_date)
        self._data = self._data[filter_boolean]

    def _load_file(self) -> None:
        row_data = []
        with open(self.path) as fid:
            file_data = csv.reader(fid)
            for r, row in enumerate(file_data):
                if r == 0:
                    continue
                elif r == 1:
                    self._header_original = row
                elif r == 2:
                    self._units_original = row
                else:
                    row_data.append(row)
        self._data = pd.DataFrame(row_data, columns=self._header_original)

    def _add_columns(self):
        def extract_ship(sampname: str):
            parts = sampname.split('-')
            return utils.map_ship(parts[0][-4:])

        def extract_serno(sampname) -> str:
            parts = sampname.split('-')
            if len(parts) == 3:
                return parts[1]
            return parts[1].split('_')[0]

        def extract_depth(sampname):
            if not sampname:
                return ''
            if '_' in sampname:
                return sampname.split('_')[-1]
            return sampname.split('-')[-1]

        def convert_timestamp(d):
            return datetime.datetime.strptime(d, '%Y-%m-%dT%H:%M:%S')

        self._data['timestamp'] = self._data['timestamp'].apply(convert_timestamp)
        self._data['date'] = self._data['timestamp'].apply(lambda x: x.date())
        self._data['year'] = self._data['date'].apply(lambda x: x.year)
        self._data['ship'] = self._data['sampleName'].apply(extract_ship)
        self._data['serno'] = self._data['sampleName'].apply(extract_serno)
        self._data['depth'] = self._data['sampleName'].apply(extract_depth)
        self._data['Rspec'] = self._data['absorbance578'].apply(float) / self._data['absorbance434'].apply(float)

    def _filter_data(self):
        def match_pattern(sampname):
            for pat in self.SAMPLE_NAME_PATTERNS:
                if pat.match(sampname):
                    return True
            return False
        boolean = self._data['action'] == 'Measure discrete'
        boolean = boolean * self._data['sampleName'].apply(match_pattern)
        self._data = self._data[boolean]

    def _reorder_columns(self):
        first_columns = ['ship', 'date', 'serno', 'depth']
        new_columns = first_columns + [col for col in self.columns if col not in first_columns]
        self._data = self._data[new_columns]


class HydrofiaTemplateData(Protocol):

    def get_data(self) -> pd.DataFrame:
        ...


class HyrdofiaExcelTemplate:
    ADDITIONAL_HEADER_MAPPING = {
        'ship': 'SHIP (correct)',
        'date': 'DATE (correct)',
        'serno': 'SERNO (correct)',
        'depth': 'DEPTH (correct)'
    }
    FILL_USER_ACTION = PatternFill(start_color='faeda2',
                                   end_color='faeda2',
                                   fill_type='solid')

    FILL_USER_ACTION_DIFFER = PatternFill(start_color='ffc996',
                                   end_color='ffc996',
                                   fill_type='solid')

    BORDER_USER_ACTION = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

    def __init__(self, path):
        self._path = pathlib.Path(path)
        self._template_create = _HyrdofiaExcelTemplateCreate(self)
        self._template_load = _HyrdofiaExcelTemplateLoad(self)

    @staticmethod
    def get_default_template_path(directory: pathlib.Path | str = None) -> pathlib.Path:
        if directory is None:
            directory = pathlib.Path(pathlib.Path.home(), 'hydrofia')
        else:
            directory = pathlib.Path(directory)
        if directory.exists() and not directory.is_dir():
            raise NotADirectoryError(directory)
        directory.mkdir(parents=True, exist_ok=True)
        return pathlib.Path(directory, f'hydrofia_mall_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')

    @property
    def path(self):
        return self._path

    def create_template(self, hydrofia: HydrofiaTemplateData, overwrite: bool = False) -> pathlib.Path:
        return self._template_create.create_template(hydrofia=hydrofia, overwrite=overwrite)

    def open_template(self):
        utils.open_file_in_default_program(self.path)

    @property
    def data(self):
        return self.get_data()

    def get_data(self):
        return self._template_load.get_data()


class _HyrdofiaExcelTemplateCreate:
    FILL_USER_ACTION = PatternFill(start_color='faeda2',
                                   end_color='faeda2',
                                   fill_type='solid')

    FILL_USER_ACTION_DIFFER = PatternFill(start_color='ffc996',
                                   end_color='ffc996',
                                   fill_type='solid')

    BORDER_USER_ACTION = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

    def __init__(self, parent: HyrdofiaExcelTemplate):
        self._parent = parent
        self.wb = Workbook()
        self.ws = self.wb.active

    @property
    def path(self):
        return self._parent.path

    @property
    def data(self):
        return self.hydrofia.get_data()

    def create_template(self, hydrofia: HydrofiaTemplateData, overwrite: bool = False) -> pathlib.Path:
        self.hydrofia = hydrofia
        if self.path.exists() and not overwrite:
            raise FileExistsError(self.path)

        self._set_info_header()
        self._set_data_header()
        self._add_data()
        self._add_correct_data()
        self._merge_cells()
        return self._save_file()

    def _set_info_header(self):
        cell = self.ws.cell(1, 1)
        cell.value = 'Gå igenom och sätt rätt värde i dessa kolumner!'
        cell.fill = HyrdofiaExcelTemplate.FILL_USER_ACTION

    def _set_data_header(self):
        """Column width help at: https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size"""
        header = list(HyrdofiaExcelTemplate.ADDITIONAL_HEADER_MAPPING.values()) + list(self.data.columns)
        for c, item in enumerate(header, 1):
            cell = self.ws.cell(2, c)
            cell.value = item
            self.ws.column_dimensions[get_column_letter(c)].width = len(item)+10
            if c <= len(HyrdofiaExcelTemplate.ADDITIONAL_HEADER_MAPPING):
                cell.fill = HyrdofiaExcelTemplate.FILL_USER_ACTION

    def _add_data(self):
        add_c = len(HyrdofiaExcelTemplate.ADDITIONAL_HEADER_MAPPING) + 1
        r = 3
        for index, series in self.data.iterrows():
            for c, value in enumerate(series, add_c):
                self.ws.cell(r, c).value = value
            r += 1

    def _add_correct_data(self):
        r = 3
        for index, series in self.data.iterrows():
            for c, key in enumerate(HyrdofiaExcelTemplate.ADDITIONAL_HEADER_MAPPING, 1):
                cell = self.ws.cell(r, c)
                cell.value = series[key]
                cell.fill = HyrdofiaExcelTemplate.FILL_USER_ACTION
                cell.border = HyrdofiaExcelTemplate.BORDER_USER_ACTION
            r += 1

    def _merge_cells(self):
        self.ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)

    def _save_file(self):
        self.wb.save(self.path)
        return self.path


class _HyrdofiaExcelTemplateLoad:

    def __init__(self, parent: HyrdofiaExcelTemplate):
        self._parent = parent
        self._data = pd.DataFrame()

    def load(self):
        self._load_template()
        self._modify_header()
        self._filter_data()
        # self._add_columns()

    def _load_template(self):
        self._data = pd.read_excel(self.path, skiprows=1, dtype=str)

    def _modify_header(self):
        for col in HyrdofiaExcelTemplate.ADDITIONAL_HEADER_MAPPING:
            self._data.drop(col, inplace=True, axis=1)
        mapper = dict((value, key) for key, value in HyrdofiaExcelTemplate.ADDITIONAL_HEADER_MAPPING.items())
        self._data.rename(columns=mapper, inplace=True)

    def _filter_data(self):
        boolean = self._data['depth'].apply(lambda x: str(x).upper()) == 'X'
        self._data = self._data[~boolean]
        self._data.reset_index(inplace=True)

    def _add_columns(self):
        self._data['year'] = self._data['date'].apply(lambda x: x.year)
        self._data['Rspec'] = self._data['absorbance578'].apply(float) / self._data['absorbance434'].apply(float)

    @property
    def path(self):
        return self._parent.path

    @property
    def data(self):
        return self._data

    def get_data(self):
        if self._data.empty:
            self.load()
        return self.data


if __name__ == '__main__':
    path = pathlib.Path(r"C:\mw\utv\HydroFIA\data\BAS 23-0147.txt")

    hf = HydrofiaExportFileDiscrete(path)

    template = HyrdofiaExcelTemplate(r'C:\mw\utv\HydroFIA/mall.xlsx')
    template.create_template(hf)
    # template.open_template()

    df = template.get_data()

