import datetime
import pathlib

import numpy as np
import pandas as pd
import inspect
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


if getattr(sys, 'frozen', False):
    ROOT_DIR = pathlib.Path(sys.executable).parent
else:
    ROOT_DIR = pathlib.Path(__file__).parent


CRM_COLOR = PatternFill(start_color='8fb7f7',
                        end_color='8fb7f7',
                        fill_type='solid')

PH_VALUE_COLOR = PatternFill(start_color='c0d6fa',
                             end_color='c0d6fa',
                             fill_type='solid')

BORDER_USER_ACTION = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))


class ExporterXlsxResultFile:
    name = 'xlsx-template-export'
    template_sheet_name = 'Rapport för utskrift'
    raw_data_sheet_name = 'Rådata'

    date_cell = [1, 9]
    signature_cell = [2, 9]
    project_cell = [6, 1]
    country_code_cell = [6, 3]
    ship_code_cell = [6, 5]
    serno_span_cell = [6, 7]

    # date_cell = [1, 12]
    # signature_cell = [2, 12]
    # project_cell = [6, 1]
    # country_code_cell = [6, 6]
    # ship_code_cell = [6, 8]
    # serno_span_cell = [6, 10]

    data_start_row = 12

    series_col = 1
    station_col = 2
    depth_col = 3
    ref_depth_col = 4
    salt_col = 5
    temp_col = 6
    ph_calc_col = 7
    # ph_col = 11
    comment_col = 8

    # series_col = 3
    # station_col = 4
    # depth_col = 6
    # ref_depth_col = 7
    # salt_col = 8
    # temp_col = 9
    # ph_calc_col = 10
    # ph_col = 11
    # comment_col = 13

    def __str__(self):
        return self.__class__.__name__

    def __init__(self, path, overwrite: bool = False):
        self._export_path = pathlib.Path(path)
        self._overwrite = overwrite
        self._template_path = pathlib.Path(ROOT_DIR, 'hydrofia_ph_template.xlsx')
        self._wb = load_workbook(self._template_path)
        self._report_ws = self._wb[self.template_sheet_name]
        self._raw_data_ws = self._wb[self.raw_data_sheet_name]
        self._data = pd.DataFrame()

    def save(self,
             data: pd.DataFrame,
             signature: str = '',
             project: str = '',
             **kwargs
             ):
        self._data = data.fillna('')
        self._write_header()
        self._write_data()
        self._write_raw_data()

        self.project = project
        self.signature = signature

        self._save_file()

    @staticmethod
    def _get_float_value(value):
        try:
            value = str(round(float(value), 3)).replace(',', '.')
        except ValueError:
            pass
        if value == 'nan':
            value = ''
        return value

    def _set_report_value(self, r: int, c: int, value: str | float, fill_color=None):
        cell = self._report_ws.cell(r, c)
        cell.value = value
        cell.alignment = Alignment(horizontal='left')
        if fill_color:
            cell.fill = fill_color

    def _set_raw_data_value(self, r: int, c: int, value: str | float):
        self._raw_data_ws.cell(r, c).value = value

    def _write_header(self):
        self.date = 'today'
        self._write_ship_code()
        self._write_country_code()
        self._write_serno_span()

    def _write_country_code(self):
        print(f"{set(self.data['country'])=}")
        values = [item for item in sorted(set(self.data['country'])) if item]
        self.country_code = ', '.join(values)

    def _write_ship_code(self):
        print(f"{set(self.data['ship'])=}")
        values = [item for item in sorted(set(self.data['ship'])) if item]
        self.ship_code = ', '.join(values)

    def _write_serno_span(self):
        # TODO: Use sorted set and see if its faster
        int_serno = [int(serno) for serno in self.data['serno'] if 'CRM' not in serno]
        # int_serno = self.data['serno'].apply(int)
        self.serno_span = f'{str(min(int_serno)).zfill(4)}-{str(max(int_serno)).zfill(4)}'

    def _write_data(self):
        r = self.data_start_row
        for index, df in self.data.groupby(['date', 'serno', 'depth']):
            s = df.iloc[-1]  # Use last replicate
            crm_color = None
            ph_color = None
            if s['calc_pH']:
                ph_color = PH_VALUE_COLOR
            if 'CRM' in s['serno']:
                crm_color = CRM_COLOR
                ph_color = CRM_COLOR
            self._set_report_value(r, self.series_col, s['serno'], fill_color=crm_color)
            self._set_report_value(r, self.station_col, s['station'], fill_color=crm_color)
            self._set_report_value(r, self.depth_col, s['depth'], fill_color=crm_color)
            self._set_report_value(r, self.ref_depth_col, self._get_float_value(s['ref_depth']), fill_color=crm_color)
            self._set_report_value(r, self.salt_col, self._get_float_value(s['salt']), fill_color=crm_color)
            self._set_report_value(r, self.temp_col, self._get_float_value(s['temp']), fill_color=crm_color)
            self._set_report_value(r, self.ph_calc_col, self._get_float_value(s['calc_pH']), fill_color=ph_color)
            # self._set_report_value(r, self.ph_col, self._get_float_value(s['pH']))
            comment = ''
            if type(s['depth']) == str and '/' in s['depth']:
                nr = s['depth'].split('/')[-1]
                comment= f'Replikat nr {nr}'
            self._set_report_value(r, self.comment_col, comment, fill_color=crm_color)
            r += 1

    def old_write_raw_data(self):
        book = load_workbook(self._export_path)
        writer = pd.ExcelWriter(self._export_path, engine='openpyxl', mode='a')
        writer.workbook = book
        self.data.to_excel(writer, sheet_name=self.raw_data_sheet_name, index=False)
        writer.close()

    def _write_raw_data(self):
        data = self.data.copy()
        data.pop('index')
        for c, name in enumerate(data, 1):
            self._set_raw_data_value(1, c, name)
        for r, row in enumerate(data.iterrows(), 2):
            for c, val in enumerate(row[1].values, 1):
                self._set_raw_data_value(r, c, val)



    def _save_file(self):
        if self._export_path.exists() and not self._overwrite:
            raise FileExistsError(self._export_path)
        self._wb.save(self._export_path)

    @property
    def data(self):
        return self._data

    @property
    def signature(self):
        return self._report_ws.cell(*self.signature_cell).value

    @signature.setter
    def signature(self, value):
        self._report_ws.cell(*self.signature_cell).value = value

    @property
    def date(self):
        return self._report_ws.cell(*self.date_cell).value

    @date.setter
    def date(self, value):
        if value.lower() == 'today':
            value = str(datetime.datetime.now().date())
        self._report_ws.cell(*self.date_cell).value = value

    @property
    def project(self):
        return self._report_ws.cell(*self.project_cell).value

    @project.setter
    def project(self, value):
        self._report_ws.cell(*self.project_cell).value = value

    @property
    def country_code(self):
        return self._report_ws.cell(*self.country_code_cell).value

    @country_code.setter
    def country_code(self, value):
        self._report_ws.cell(*self.country_code_cell).value = value

    @property
    def ship_code(self):
        return self._report_ws.cell(*self.ship_code_cell).value

    @ship_code.setter
    def ship_code(self, value):
        self._report_ws.cell(*self.ship_code_cell).value = value

    @property
    def serno_span(self):
        return self._report_ws.cell(*self.serno_span_cell).value

    @serno_span.setter
    def serno_span(self, value):
        self._report_ws.cell(*self.serno_span_cell).value = value


class ExporterTxt:
    name = 'txt-export'

    def __str__(self):
        return self.__class__.__name__

    def __init__(self, path: pathlib.Path | str, overwrite: bool = False):
        self.path = pathlib.Path(path)
        self._overwrite = overwrite

    def save(self, data: pd.DataFrame, **kwargs):
        if self.path.exists() and not self._overwrite:
            raise FileExistsError(self.path)
        leading_cols = ['year', 'ship', 'date', 'serno', 'depth', 'calc_pH', 'salt', 'temp', 'ref_depth', 'Rspec']
        other_cols = [col for col in data.columns if col not in leading_cols]
        new_columns = leading_cols + other_cols
        new_data = data[new_columns]
        if 'index' in new_data.columns:
            new_data.pop('index')
        new_data.to_csv(self.path, sep='\t', index=False)


class Exporters:
    exporters = [
        ExporterTxt,
        ExporterXlsxResultFile
    ]

    @staticmethod
    def get_arguments_from_class(cls):
        return list(inspect.signature(cls).parameters)

    def __init__(self):
        self._exporters = {}
        for cls in Exporters.exporters:
            info = dict(
                cls=cls,
                name=cls.name,
                args=Exporters.get_arguments_from_class(cls)
            )
            self._exporters[info['name']] = info

    def get_exporters(self):
        return self._exporters

    def get_exporter_list(self):
        return list(self._exporters)

    def get_exporter_cls(self, name):
        return self._exporters[name]['cls']

    def get_exporter_args(self, name):
        return self._exporters[name]['args']


if __name__ == '__main__':
    path = pathlib.Path(ROOT_DIR, 'hydrofia_ph_template.xlsx')
    save_path = pathlib.Path(ROOT_DIR, 'hydrofia_ph_template_out.xlsx')
    print(path.exists())
    wb = load_workbook(path)
    ws = wb['Rapport för utskrift']

    e = ExporterXlsxResultFile()
    e.signature = 'MWen'
    e.date = 'today'
    e.country_code = 'SE'
    e.ship_code = '77'

    e.save(save_path)

